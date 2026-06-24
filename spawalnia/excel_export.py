from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

C_HEADER = 'FF1A5276'
C_SUBHEAD = 'FF1A6E8A'
C_OK     = 'FF1E8449'
C_NG     = 'FFC0392B'
C_GRAY   = 'FFF2F3F4'
C_WHITE  = 'FFFFFFFF'
FMT_NUM  = '0.00'


def _border():
    s = Side(style='thin', color='FFB2BABB')
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(records, zo_filter='') -> BytesIO:
    if not HAS_OPENPYXL:
        raise RuntimeError('Biblioteka openpyxl nie jest zainstalowana.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Spawalnia'

    col_widths = [18, 14, 14, 14, 12, 12, 12, 16, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def cell(row, col, value=None, bold=False, bg=None, fg='FF000000',
             fmt=None, align='center'):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=10)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if fmt:
            c.number_format = fmt
        c.border = _border()
        return c

    # ── Nagłówek dokumentu ────────────────────────────────────────────────────
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = 'KONTROLA SPAWALNI'
    title_cell.font = Font(bold=True, color='FFFFFFFF', size=13)
    title_cell.fill = PatternFill('solid', fgColor=C_HEADER)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:F2')
    sub = ws['A2']
    sub.value = f'Filtr ZO: {zo_filter}' if zo_filter else 'Wszystkie rekordy'
    sub.font = Font(color='FFFFFFFF', size=9)
    sub.fill = PatternFill('solid', fgColor=C_SUBHEAD)
    sub.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('G2:K2')
    date_cell = ws['G2']
    date_cell.value = datetime.now().strftime('%d.%m.%Y %H:%M')
    date_cell.font = Font(color='FFFFFFFF', size=9)
    date_cell.fill = PatternFill('solid', fgColor=C_SUBHEAD)
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Nagłówki kolumn ───────────────────────────────────────────────────────
    headers = ['NR ZO', 'OTWOROWANIE', 'PRZEKĄTNA', 'ODCHYŁKA [mm]',
               'POMIAR 1 [cm]', 'POMIAR 2 [cm]', 'POMIAR 3 [cm]',
               'JAKOŚĆ WYCIĘCIA', 'SPAWACZ', 'GIĘCIE', 'CIĘCIE', 'DATA']
    for col, h in enumerate(headers, 1):
        cell(4, col, h, bold=True, bg=C_SUBHEAD, fg='FFFFFFFF')
    ws.row_dimensions[4].height = 30

    # ── Dane ──────────────────────────────────────────────────────────────────
    for i, rec in enumerate(records):
        row = 5 + i
        bg = C_WHITE if i % 2 == 0 else C_GRAY

        def ok_ng_fg(val):
            if val == 'OK':
                return C_OK
            if val == 'NG':
                return C_NG
            return 'FF555555'

        cell(row, 1, rec.zo_number, bold=True, bg=bg, align='left')
        c2 = cell(row, 2, rec.otworowanie or '—', bg=bg, fg=ok_ng_fg(rec.otworowanie))
        c2.font = Font(bold=True, color=ok_ng_fg(rec.otworowanie), size=10)
        c3 = cell(row, 3, rec.przekatna or '—', bg=bg, fg=ok_ng_fg(rec.przekatna))
        c3.font = Font(bold=True, color=ok_ng_fg(rec.przekatna), size=10)

        def fmt_float(v):
            return round(v, 2) if v is not None else None

        cell(row, 4, fmt_float(rec.przekatna_odchylka), bg=bg, fmt=FMT_NUM)
        cell(row, 5, fmt_float(rec.pomiar1), bg=bg, fmt=FMT_NUM)
        cell(row, 6, fmt_float(rec.pomiar2), bg=bg, fmt=FMT_NUM)
        cell(row, 7, fmt_float(rec.pomiar3), bg=bg, fmt=FMT_NUM)

        c8 = cell(row, 8, rec.jakosc_wyciecia or '—', bg=bg, fg=ok_ng_fg(rec.jakosc_wyciecia))
        c8.font = Font(bold=True, color=ok_ng_fg(rec.jakosc_wyciecia), size=10)

        op_label     = rec.operator.initials        if rec.operator        else '—'
        giecie_label = rec.giecie_operator.initials if rec.giecie_operator else '—'
        ciecie_label = rec.ciecie_operator.initials if rec.ciecie_operator else '—'
        cell(row, 9,  op_label,     bg=bg)
        cell(row, 10, giecie_label, bg=bg)
        cell(row, 11, ciecie_label, bg=bg)
        date_label = rec.created_at.strftime('%d.%m.%Y') if rec.created_at else '—'
        cell(row, 12, date_label, bg=bg)

    # ── Stopka ────────────────────────────────────────────────────────────────
    footer_row = 5 + len(records) + 1
    ws.merge_cells(f'A{footer_row}:L{footer_row}')
    fc = ws[f'A{footer_row}']
    fc.value = (f'Wygenerowano: {datetime.now().strftime("%d.%m.%Y %H:%M")}  |  '
                f'Liczba rekordów: {len(records)}')
    fc.font = Font(italic=True, color='FF888888', size=8)
    fc.alignment = Alignment(horizontal='left', vertical='center')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

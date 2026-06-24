"""
Generator pliku Excel dla kosztorysu — format zbliżony do wzoru PSH.
Wymaga: openpyxl
"""
from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Kolory
C_HEADER  = 'FF1A5276'   # granat (Kubiak)
C_SUBHEAD = 'FF1A6E8A'   # teal
C_TOTAL   = 'FFEAF2F8'   # jasny niebieski
C_WHITE   = 'FFFFFFFF'
C_GRAY    = 'FFF2F3F4'

FMT_PLN   = '#,##0.00 "PLN"'
FMT_KG    = '#,##0.00 "kg"'
FMT_M2    = '#,##0.00 "m²"'
FMT_PCT   = '0.00"%"'
FMT_NUM   = '#,##0.00'


def _hdr_style(bold=True, color=C_HEADER, font_color='FFFFFFFF', size=10):
    return {
        'font': Font(bold=bold, color=font_color, size=size),
        'fill': PatternFill('solid', fgColor=color),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }


def _apply(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def _border(thin=True):
    s = Side(style='thin' if thin else 'hair', color='FFB2BABB')
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(quote) -> BytesIO:
    if not HAS_OPENPYXL:
        raise RuntimeError('Biblioteka openpyxl nie jest zainstalowana. '
                           'Uruchom: pip install openpyxl')

    cfg  = quote.config
    calc = cfg.calculation or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kosztorys'

    # ── Szerokości kolumn ───────────────────────────────────────────────────
    col_widths = [28, 6, 6, 8, 8, 8, 7, 7, 9, 9, 11, 11, 12, 10, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def w(row, col, value=None, bold=False, bg=None, fg='FF000000',
          fmt=None, align='left', border=True, size=10, wrap=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, color=fg, size=size)
        if bg:
            cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fmt:
            cell.number_format = fmt
        if border:
            cell.border = _border()
        return cell

    def merge_write(row, c1, c2, value=None, bold=False, bg=None, fg='FF000000',
                    align='left', size=10):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=value)
        cell.font = Font(bold=bold, color=fg, size=size)
        if bg:
            cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        return cell

    # ── Nagłówek dokumentu ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    merge_write(1, 1, 10, f'KOSZTORYS  {quote.number}  —  {quote.cabinet_type.name}',
                bold=True, bg=C_HEADER, fg=C_WHITE, align='center', size=13)
    merge_write(1, 11, 15, f'Data: {datetime.now().strftime("%d.%m.%Y")}',
                bold=True, bg=C_HEADER, fg=C_WHITE, align='right', size=11)

    ws.row_dimensions[2].height = 20
    merge_write(2, 1, 5, f'Klient: {quote.client_name}', bold=True, size=11)
    merge_write(2, 6, 10, f'Status: {quote.status_label}', size=10)
    if quote.notes:
        merge_write(2, 11, 15, f'Uwagi: {quote.notes}', size=9)

    # ── Dane wejściowe ──────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    merge_write(3, 1, 3, f'{quote.cabinet_type.code}', bold=True, bg=C_SUBHEAD, fg=C_WHITE, size=11)
    w(3, 2, 'szer.', bold=True, bg=C_SUBHEAD, fg=C_WHITE, align='center')
    w(3, 3, 'wys.', bold=True, bg=C_SUBHEAD, fg=C_WHITE, align='center')
    w(3, 4, 'głęb.', bold=True, bg=C_SUBHEAD, fg=C_WHITE, align='center')
    w(3, 6, 'Objętość [dm³]', bold=True, bg=C_GRAY, align='center')

    w(4, 2, cfg.width, align='center', bg=C_GRAY)
    w(4, 3, cfg.height, align='center', bg=C_GRAY)
    w(4, 4, cfg.depth, align='center', bg=C_GRAY)
    w(4, 6, int(calc.get('volume', 0)), align='center', bold=True)

    # Opcje
    opts = [
        ('Grubość obudowy', cfg.thickness_body, 'mm', 5),
        ('Grubość płyty',   cfg.thickness_plate, 'mm', 6),
        ('Płyta montażowa', 1 if cfg.has_mounting_plate else 0, '1=tak', 7),
        ('Tył spawany',     1 if cfg.back_welded else 0,        '1=tak', 8),
        ('Drzwi pojed.',    1 if cfg.door_single else 0,        '1=tak', 9),
        ('Drzwi podwójne',  1 if cfg.door_double else 0,        '1=tak', 10),
        ('Zam. trzypunkt.', 1 if cfg.lock_three_point else 0,   '1=tak', 11),
        ('Kapa (szt)',       cfg.cable_entries,                   'szt',  12),
        ('Cokół',            1 if cfg.plinth else 0,             '1=tak', 13),
        ('Kolor niestand.',  1 if cfg.non_standard_color else 0, '1=tak', 14),
    ]
    for i, (name, val, unit, row) in enumerate(opts):
        w(row, 1, name, size=9)
        w(row, 2, val, align='center', bold=True)
        w(row, 3, unit, size=9, align='center')

    # ── Tabela elementów blaszanych ─────────────────────────────────────────
    HDR_ROW = 16
    ws.row_dimensions[HDR_ROW].height = 36

    hdr = ['Element', 'Il.', 'j.m.', 'Mat.', 'Dł', 'Szer', 'Pow.m²',
           'Gr.mm', 'Waga/szt', 'Waga łącz.', 'Blacha', 'Malow.',
           'Koszt/szt', 'Robocizna', 'Koszt łącznie']
    for col, h in enumerate(hdr, 1):
        cell = w(HDR_ROW, col, h, bold=True, bg=C_SUBHEAD, fg=C_WHITE,
                 align='center', wrap=True)

    DATA_ROW = HDR_ROW + 1
    elements = calc.get('elements', [])
    row = DATA_ROW

    for e in elements:
        if e.get('qty', 0) == 0:
            continue
        ws.row_dimensions[row].height = 16
        bg = C_GRAY if row % 2 == 0 else None
        w(row, 1, e['name'], bg=bg)
        w(row, 2, e['qty'], align='center', bg=bg)
        w(row, 3, e['unit'], align='center', bg=bg)
        w(row, 4, e['material'], align='center', bg=bg)
        w(row, 5, e['L'], align='right', bg=bg)
        w(row, 6, e['W'], align='right', bg=bg)
        w(row, 7, e['area'], fmt=FMT_NUM, align='right', bg=bg)
        w(row, 8, e['thickness'], align='right', bg=bg)
        w(row, 9, e['weight_per'], fmt=FMT_NUM, align='right', bg=bg)
        w(row, 10, e['weight_total'], fmt=FMT_NUM, align='right', bg=bg)
        w(row, 11, e['cost_sheet'], fmt=FMT_PLN, align='right', bg=bg)
        w(row, 12, e['cost_paint'] if e['cost_paint'] else None, fmt=FMT_PLN, align='right', bg=bg)
        w(row, 13, e['cost_per'], fmt=FMT_PLN, align='right', bg=bg)
        w(row, 15, e['cost_total'], fmt=FMT_PLN, align='right', bold=True, bg=bg)
        row += 1

    # Odpad
    waste = calc.get('waste', {})
    if waste:
        ws.row_dimensions[row].height = 16
        w(row, 1, waste.get('name', 'Odpad 15%'), bg=C_GRAY)
        w(row, 2, 1, align='center', bg=C_GRAY)
        w(row, 10, waste.get('weight_total'), fmt=FMT_NUM, align='right', bg=C_GRAY)
        w(row, 13, waste.get('cost_total'), fmt=FMT_PLN, align='right', bg=C_GRAY)
        w(row, 15, waste.get('cost_total'), fmt=FMT_PLN, align='right', bold=True, bg=C_GRAY)
        row += 1

    # ── Osprzęt ─────────────────────────────────────────────────────────────
    row += 1
    merge_write(row, 1, 15, 'OSPRZĘT I MATERIAŁY POMOCNICZE',
                bold=True, bg=C_SUBHEAD, fg=C_WHITE)
    row += 1

    for h in calc.get('hardware', []):
        bg = C_GRAY if row % 2 == 0 else None
        w(row, 1, h['name'], bg=bg)
        w(row, 2, h['qty'], align='center', bg=bg)
        w(row, 3, h['unit'], align='center', bg=bg)
        w(row, 13, h['unit_price'], fmt=FMT_PLN, align='right', bg=bg)
        w(row, 15, h['total'], fmt=FMT_PLN, align='right', bold=True, bg=bg)
        row += 1

    # ── Robocizna i usługi ───────────────────────────────────────────────────
    row += 1
    merge_write(row, 1, 15, 'ROBOCIZNA I USŁUGI',
                bold=True, bg=C_SUBHEAD, fg=C_WHITE)
    row += 1

    for s in calc.get('services', []):
        bg = C_GRAY if row % 2 == 0 else None
        w(row, 1, s['name'], bg=bg)
        w(row, 2, s['qty'], align='center', bg=bg)
        w(row, 3, s['unit'], align='center', bg=bg)
        w(row, 13, s['unit_price'], fmt=FMT_PLN, align='right', bg=bg)
        w(row, 15, s['total'], fmt=FMT_PLN, align='right', bold=True, bg=bg)
        row += 1

    # ── Podsumowanie ─────────────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 24
    merge_write(row, 1, 13, 'KOSZT WŁASNY', bold=True, bg=C_TOTAL, align='right', size=12)
    cell = w(row, 14, None, bold=True, bg=C_TOTAL)
    merge_write(row, 14, 15, calc.get('cost_total'), bold=True, bg=C_TOTAL,
                align='right', size=13)
    ws.cell(row=row, column=14).number_format = FMT_PLN
    row += 1

    rows_summary = [
        ('Marża (mnożnik)',   calc.get('margin')),
        (f'Cena katalogowa',  calc.get('price_catalog')),
        (f'Rabat {calc.get("discount_pct", 0)}%', calc.get('price_discount')),
        (f'Bonus {calc.get("bonus_pct", 0)}%', calc.get('price_bonus')),
        ('Zyskowność',        f'{calc.get("profitability", 0):.2f}%'),
    ]
    for label, value in rows_summary:
        ws.row_dimensions[row].height = 20
        merge_write(row, 1, 13, label, bold=True, align='right')
        cell = ws.cell(row=row, column=14, value=value)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right')
        if isinstance(value, float):
            cell.number_format = FMT_PLN
        ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=15)
        row += 1

    # ── Stopka ──────────────────────────────────────────────────────────────
    row += 1
    merge_write(row, 1, 15,
                f'Wycena wygenerowana: {datetime.now().strftime("%d.%m.%Y %H:%M")}  |  '
                f'Autor: {quote.created_by.username}',
                size=9, align='center')

    # Zamroź górne wiersze
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
